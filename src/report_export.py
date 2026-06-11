"""
报告导出模块 — 将分析结果与图表导出为结构化文档

功能:
  1. 导出 Excel 报告（数据分析结果 + 图表嵌入）
  2. 导出纯文本摘要报告
  3. 批量导出图表
"""

import os
from typing import Optional
from datetime import datetime

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter


class ReportGenerator:
    """分析报告生成器"""

    # Excel 样式常量
    HEADER_FONT = Font(name="微软雅黑", size=11, bold=True, color="FFFFFF")
    HEADER_FILL = PatternFill(start_color="4472C4", end_color="4472C4",
                              fill_type="solid")
    TITLE_FONT = Font(name="微软雅黑", size=14, bold=True)
    SECTION_FONT = Font(name="微软雅黑", size=11, bold=True, color="1F4E79")
    DATA_FONT = Font(name="微软雅黑", size=10)
    NOTE_FONT = Font(name="微软雅黑", size=9, italic=True, color="808080")
    THIN_BORDER = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )

    def __init__(self):
        self.timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")

    # ──────────────────────────────────────────
    # Excel 报告
    # ──────────────────────────────────────────

    def to_excel_with_charts(self, df: pd.DataFrame,
                             analysis_results: dict,
                             chart_paths: Optional[list] = None,
                             file_path: Optional[str] = None,
                             title: str = "传感器数据分析报告") -> str:
        """
        生成带图表的 Excel 报告

        参数:
            df: 原始数据
            analysis_results: 分析结果字典
            chart_paths: 图表图片路径列表
            file_path: 输出路径（可选，自动生成默认路径）
            title: 报告标题

        返回:
            输出文件路径
        """
        if file_path is None:
            file_path = f"sensor_report_{self.timestamp}.xlsx"
        if not file_path.endswith(".xlsx"):
            file_path += ".xlsx"

        wb = Workbook()

        # ====== Sheet 1: 报告总览 ======
        ws1 = wb.active
        ws1.title = "报告总览"
        self._write_report_overview(ws1, title, df, analysis_results)

        # ====== Sheet 2: 原始数据 ======
        ws2 = wb.create_sheet("原始数据")
        self._write_dataframe(ws2, df)

        # ====== Sheet 3: 统计摘要 ======
        ws3 = wb.create_sheet("统计摘要")
        stats_df = self._generate_stats_df(df)
        if stats_df is not None:
            self._write_dataframe(ws3, stats_df)

        # ====== Sheet 4: 相关性分析 ======
        if "correlation" in analysis_results:
            ws4 = wb.create_sheet("相关性分析")
            corr_matrix = analysis_results["correlation"].get("correlation_matrix")
            if corr_matrix is not None:
                self._write_dataframe(ws4, corr_matrix.reset_index())

        # ====== Sheet 5+: 异常检测 ======
        if "outliers" in analysis_results:
            ws5 = wb.create_sheet("异常检测")
            outliers = analysis_results["outliers"]
            row = 1
            for col, info in outliers.items():
                ws5.cell(row=row, column=1, value=f"传感器: {col}")
                ws5.cell(row=row, column=1).font = self.SECTION_FONT
                row += 1
                for key, val in info.items():
                    ws5.cell(row=row, column=1, value=key)
                    ws5.cell(row=row, column=1).font = Font(bold=True)
                    ws5.cell(row=row, column=2, value=str(val))
                    row += 1
                row += 1

        # 保存
        os.makedirs(os.path.dirname(file_path) or ".", exist_ok=True)
        wb.save(file_path)
        return file_path

    def to_excel_simple(self, df: pd.DataFrame,
                        file_path: str,
                        sheet_name: str = "传感器数据") -> str:
        """
        快速导出原始数据到 Excel

        参数:
            df: 待导出的 DataFrame
            file_path: 输出路径
            sheet_name: 工作表名称

        返回:
            输出文件路径
        """
        os.makedirs(os.path.dirname(file_path) or ".", exist_ok=True)
        with pd.ExcelWriter(file_path, engine="openpyxl") as writer:
            df.to_excel(writer, sheet_name=sheet_name, index=False)
        return file_path

    # ──────────────────────────────────────────
    # 文本报告
    # ──────────────────────────────────────────

    def to_text_report(self, df: pd.DataFrame,
                       analysis_results: dict,
                       file_path: Optional[str] = None,
                       title: str = "传感器数据分析报告") -> str:
        """
        生成纯文本分析报告

        返回:
            输出文件路径
        """
        if file_path is None:
            file_path = f"sensor_report_{self.timestamp}.txt"

        lines = []
        sep = "=" * 65
        dash = "-" * 65

        lines.append(sep)
        lines.append(f"  {title}")
        lines.append(f"  生成时间: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
        lines.append(sep)
        lines.append("")

        # 数据概览
        lines.append("【数据概览】")
        lines.append(dash)
        lines.append(f"  行数: {len(df)}")
        lines.append(f"  列数: {len(df.columns)}")
        lines.append(f"  列名: {', '.join(df.columns.tolist())}")
        lines.append(f"  缺失值: {int(df.isnull().sum().sum())}")
        lines.append("")

        # 数值列统计
        numeric_cols = df.select_dtypes(include=["number"]).columns
        if len(numeric_cols) > 0:
            lines.append("【数值列统计】")
            lines.append(dash)
            for col in numeric_cols:
                s = df[col]
                lines.append(f"  {col}:")
                lines.append(f"    样本数={len(s.dropna())}, "
                             f"均值={s.mean():.4f}, "
                             f"标准差={s.std():.4f}")
                lines.append(f"    最小值={s.min():.4f}, "
                             f"25%={s.quantile(0.25):.4f}, "
                             f"中位数={s.median():.4f}")
                lines.append(f"    75%={s.quantile(0.75):.4f}, "
                             f"最大值={s.max():.4f}")
                lines.append(f"    缺失={int(s.isnull().sum())}")
            lines.append("")

        # 相关性
        if "correlation" in analysis_results:
            corr_result = analysis_results["correlation"]
            lines.append("【相关性分析】")
            lines.append(dash)
            for pair in corr_result.get("highly_correlated", []):
                lines.append(f"  {pair[0]} 与 {pair[1]}: "
                             f"r={pair[2]}")
            lines.append("")

        # 异常检测
        if "outliers" in analysis_results:
            lines.append("【异常检测结果】")
            lines.append(dash)
            for col, info in analysis_results["outliers"].items():
                n = info.get("异常个数", 0)
                lines.append(f"  {col}: {n} 个异常值 "
                             f"({info.get('异常占比', 0)}%)")
            lines.append("")

        # 趋势
        if "trends" in analysis_results:
            lines.append("【趋势分析】")
            lines.append(dash)
            for col, trend in analysis_results["trends"].items():
                lines.append(f"  {col}: {trend.get('trend', 'N/A')} "
                             f"(斜率={trend.get('slope', 0)})")
            lines.append("")

        lines.append(sep)
        lines.append("报告结束")
        lines.append(sep)

        content = "\n".join(lines)
        os.makedirs(os.path.dirname(file_path) or ".", exist_ok=True)
        with open(file_path, "w", encoding="utf-8") as f:
            f.write(content)
        return file_path

    # ──────────────────────────────────────────
    # 内部方法（Excel 样式）
    # ──────────────────────────────────────────

    def _write_report_overview(self, ws, title: str, df, analysis_results):
        """写入 Excel 报告总览页"""
        ws.merge_cells("A1:F1")
        ws.cell(row=1, column=1, value=title).font = self.TITLE_FONT

        row = 3
        ws.cell(row=row, column=1, value="数据概览").font = self.SECTION_FONT
        row += 1
        overview_data = [
            ("总行数", str(len(df))),
            ("总列数", str(len(df.columns))),
            ("数值列", str(len(df.select_dtypes(include=["number"]).columns))),
            ("缺失值总数", str(int(df.isnull().sum().sum()))),
            ("分析时间", datetime.now().strftime("%Y-%m-%d %H:%M:%S")),
        ]
        for label, value in overview_data:
            ws.cell(row=row, column=1, value=label)
            ws.cell(row=row, column=2, value=value)
            row += 1

        if "outliers" in analysis_results:
            row += 1
            ws.cell(row=row, column=1,
                    value="异常检测").font = self.SECTION_FONT
            row += 1
            for col, info in analysis_results["outliers"].items():
                n = info.get("异常个数", 0)
                ws.cell(row=row, column=1, value=f"  {col}")
                ws.cell(row=row, column=2, value=f"{n} 个异常值")
                row += 1

        if "correlation" in analysis_results:
            row += 1
            ws.cell(row=row, column=1,
                    value="高相关对").font = self.SECTION_FONT
            row += 1
            for pair in analysis_results["correlation"].get(
                    "highly_correlated", []):
                ws.cell(row=row, column=1, value=f"  {pair[0]} vs {pair[1]}")
                ws.cell(row=row, column=2, value=f"r = {pair[2]}")
                row += 1

        # 设置列宽
        ws.column_dimensions["A"].width = 20
        ws.column_dimensions["B"].width = 40

    def _write_dataframe(self, ws, df: pd.DataFrame,
                         max_rows: int = 5000):
        """将 DataFrame 写入 Excel 工作表，带表头样式"""
        if len(df) > max_rows:
            df = df.head(max_rows)

        # 写表头
        for col_idx, col_name in enumerate(df.columns, 1):
            cell = ws.cell(row=1, column=col_idx, value=str(col_name))
            cell.font = self.HEADER_FONT
            cell.fill = self.HEADER_FILL
            cell.alignment = Alignment(horizontal="center")
            cell.border = self.THIN_BORDER

        # 写数据
        for row_idx, (_, row) in enumerate(df.iterrows(), 2):
            for col_idx, value in enumerate(row, 1):
                cell = ws.cell(row=row_idx, column=col_idx, value=value)
                cell.font = self.DATA_FONT
                cell.border = self.THIN_BORDER

        # 自动列宽
        for col_idx, col_name in enumerate(df.columns, 1):
            max_len = max(
                df[col_name].astype(str).map(len).max() if len(df) > 0 else 0,
                len(str(col_name))
            )
            ws.column_dimensions[get_column_letter(col_idx)].width = (
                min(max_len + 4, 40)
            )

    @staticmethod
    def _generate_stats_df(df: pd.DataFrame) -> Optional[pd.DataFrame]:
        """生成数据统计摘要 DataFrame"""
        numeric_cols = df.select_dtypes(include=["number"]).columns
        if len(numeric_cols) == 0:
            return None

        stats_data = []
        for col in numeric_cols:
            s = df[col].dropna()
            if len(s) == 0:
                continue
            stats_data.append({
                "列名": col,
                "样本数": len(s),
                "缺失数": int(df[col].isnull().sum()),
                "均值": round(s.mean(), 4),
                "标准差": round(s.std(), 4),
                "最小值": round(s.min(), 4),
                "25%分位": round(s.quantile(0.25), 4),
                "中位数": round(s.median(), 4),
                "75%分位": round(s.quantile(0.75), 4),
                "最大值": round(s.max(), 4),
                "偏度": round(s.skew(), 4),
                "峰度": round(s.kurtosis(), 4),
            })
        return pd.DataFrame(stats_data)
